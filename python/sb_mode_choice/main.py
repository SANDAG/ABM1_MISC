import openpyxl
import pandas as pd
import pyodbc
import settings
import time


# start execution timer
start = time.time()

# load workbook template
template = openpyxl.load_workbook(settings.template_path)

# initialize output Report based on template
templateWriter = pd.ExcelWriter(
    path=settings.template_path,
    mode="w",
    engine="openpyxl")

templateWriter.book = template
templateWriter.sheets = dict((ws.title, ws) for ws in template.worksheets)

# build SQL Server connection string
conn = pyodbc.connect("DRIVER={SQL Server Native Client 11.0};SERVER="
                      + settings.server +
                      ";DATABASE=" +
                      settings.database +
                      ";Trusted_Connection=yes;")

# write user inputs to the ModeChoice report Input sheet
reportInput = template['Input']

# user inputs ABM database scenario_id
scenario_id = input("Enter an ABM [scenario_id]: ")
reportInput['C7'] = scenario_id

# user inputs project info, e.g. NAVWAR Redevelopment
reportInput['C5'] = input("Enter project name: ")

# user inputs scenario header, e.g. Low Density with Transit Center
reportInput['C8'] = input("Enter  Scenario Header: ")

# user inputs report geography, e.g. Project Site
reportInput['C9'] = input("Enter Report Geography: ")

# user inputs report filename suffix
file_suffix = input("Enter report file name suffix: ")

# delete previous mgra list from report Input sheet
reportInput.move_range("E2:F2", rows=0, cols=12)  # move header
reportInput.delete_cols(5, 5)  # delete columns
reportInput.move_range("L2:M2", rows=0, cols=-7)  # move header back

print("Beginning [scenario_id]: " + scenario_id)
scenario_id = int(scenario_id)

# for each dictionary of ad-hoc queries ------------------
for qry in settings.adhoc_queries:

    # execute ad-hoc query
    result = pd.read_sql_query(
        sql=qry["query"],
        con=conn
    )

    # write to specified Excel template sheet
    result.to_excel(
        excel_writer=templateWriter,
        sheet_name=qry["sheet"],
        na_rep="NULL",
        header=True,
        index=False,
        startrow=qry["row"],
        startcol=qry["column"],
        engine="openpyxl")


print("Running: Report queries")

# for each sheet of report queries --------------------
for qry in settings.report_queries:

    print("Running: " + qry["sp"] + " " + qry["sheet"])

    # execute stored procedure with specified arguments
    result = pd.read_sql_query(
        sql="EXECUTE " + qry["sp"] + " " + qry["args"],
        con=conn,
        params=[scenario_id]
    )

    # validate scenario id (check for hard-coded id in stored procedure)
    for s in result["scenario_id"]:
        if s != scenario_id:
            msg = ("Invalid scenario id hard-coded:",s,"should be:", scenario_id, "for sheet:", qry["sheet"])
            print(result)
            raise ValueError(msg)

    # set first row of sql query to be column headers
    new_header = result.iloc[0]  # first row for column names
    result = result[1:]  # data less the first row
    result.columns = new_header  # set the column names

    # set columns to numeric where possible
    result = result.apply(pd.to_numeric) # convert columns to numeric

    # delete rows from template that have data from previous run
    worksheet = templateWriter.sheets[qry["sheet"]]
    worksheet.delete_rows(2, worksheet.max_row-1)

    # write results to specified Excel template sheet
    result.to_excel(
        excel_writer=templateWriter,
        sheet_name=qry["sheet"],
        na_rep="NULL",
        header=True,
        index=False,
        startrow=1,
        engine="openpyxl")

    # align cells for pretty output
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')


print("Elapsed time: %5.2f minutes" % ((time.time() - start) / 60.0))

# if protection option set
if settings.template_protect is True:
    # for each worksheet in the template
    for sheet in template.worksheets:
        # set protection and password
        sheet.protection.set_password(settings.template_password)

# save the completed template
template.save(settings.report_write_path + " - " + file_suffix + ".xlsx")
